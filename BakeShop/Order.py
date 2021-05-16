#--------------------FIT5136_SEM_02_Team 09  Assignment 4B - Bakeshop System -------------------------------
# Author Names & IDs : Dipal Vyas(31847390) , Aashna Jain(29802911) , Nivedha Balasubramanian(30729963), Kanika Bansal(31030319)
# Start Date : 16th September, 2020
# End Date : 1st November, 2020
#------------------------------------------------------------------------------------------------------------
# Class Description: Order class is for reading data from 'order' excel sheet for getting value of last order id and
#                    storing order details in 'order' sheet when order is placed.

# Main functions: 1.read_from_excel 2. write_inventory
#Getter Methods : Mainly used For getting values for order id, staff id.
#setter Methods: Used for creating string of item name and item quantity for storing in the same format as in given database.

#-------------------Importing libraries and clsses-------------------------
import pandas as pd
from Employee import Employee
import os
import openpyxl
from Item import Item
from datetime import datetime
#---------------------------------------------------------------------------

class Order():
    def __init__(self):
        self.item_detail = []
        self.item_to_store = ''
        self.list_of_items = []
        self.list_of_quantities = []
        self.list_of_costs = []
        self.order_total = 0
        self.quantities_to_store = ''
        self.cost_to_store = ''
        self.input_order_date = ''
        self.cust_contact_no = ''

    def get_date_time(self):
        date_time = datetime.now()
        date = date_time.strftime("%m/%d/%Y")
        time = date_time.strftime("%H:%M:%S")
        return date, time

    # Following function is created to get last order id from 'orders' excel sheet to set new order id while placing new order.
    def get_order_id(self):
        df_orders = self.read_from_excel()
        data =df_orders["order id"]
        max_order_id =data.max()
        return max_order_id + 1

    def get_list_of_cost(self):
        return self.list_of_costs

    def get_list_of_quantity(self):
        return self.list_of_quantities

    def get_list_of_item(self):
        return self.list_of_items

    def get_item_to_store(self):
        return self.item_to_store

    def get_quantity_to_store(self):
        return self.quantities_to_store

    def get_cost_to_store(self):
        return self.cost_to_store

    def set_order_store_id(self, user):
        # Returns Store ID string if only single store or a list of store IDs if multiple stores under an employee
        employee: Employee = Employee()
        store_id = employee.get_emp_store_id(user)
        return store_id

    def set_order_staff_id(self, user):
        employee: Employee = Employee()
        staff_id = employee.get_emp_id(user)
        return staff_id

    def set_item_to_store(self,input_item_no):
        # creating string in specific format to write item in excel file
        if self.item_to_store == '':
            self.item_to_store = str(self.item_detail[input_item_no - 1][1])
        else:
            self.item_to_store = self.item_to_store + ', ' + str(self.item_detail[input_item_no - 1][1])
        self.list_of_items.append(self.item_detail[input_item_no - 1][1])

    def set_quantity_to_store(self, input_quantity):
        # creating string in specific format to write quantity in excel file
        if self.quantities_to_store == '':
            self.quantities_to_store = str(input_quantity)
        else:
            self.quantities_to_store = self.quantities_to_store + ', ' + str(input_quantity)
        self.list_of_quantities.append(input_quantity)

    def set_cost_to_store(self, item_total):
        # creating string in specific format to write cost in excel file
        if self.cost_to_store == '':
            self.cost_to_store = str(item_total)
        else:
            self.cost_to_store = self.cost_to_store + '|' + str(item_total)
        self.list_of_costs.append(item_total)

    def read_from_excel(self):
        path = os.getcwd().replace("\\", "/")
        data = pd.read_excel(path + '/SampleData.xlsx', sheet_name='orders')
        df = pd.DataFrame(data)
        return df

    # Storing Order details in 'orders' worksheet after creating order.
    def write_order(self,storeID,orderID,staffID,custName,orderStatus,custNo):
        try:
            date, time = self.get_date_time()
            row1 = [storeID, orderID, staffID, self.get_item_to_store(), self.quantities_to_store, self.cost_to_store, str(self.order_total), date, time, custName, orderStatus, custNo]
            path = os.getcwd().replace("\\", "/")
            file = path + '/SampleData.xlsx'
            wb = openpyxl.load_workbook(filename=file)
            ws = wb['orders']
            row = ws.max_row + 1
            for col, entry in enumerate(row1, start=1):
                ws.cell(row=row, column=col, value=entry)

            wb.save(file)
        except PermissionError:
            print("Permission to excel file denied. Please ensure that the file is closed and run the program again")
            exit(0)

