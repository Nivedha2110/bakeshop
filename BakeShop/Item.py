#--------------------FIT5136_SEM_02_Team 09  Assignment 4B - Bakeshop System -------------------------------
# Author Names & IDs : Dipal Vyas(31847390) , Aashna Jain(29802911) , Nivedha Balasubramanian(30729963), Kanika Bansal(31030319)
# Start Date : 16th September , 2020
# End Date : 1st November, 2020
#------------------------------------------------------------------------------------------------------------
# Class Description: Item class is used for getting item details from 'item' escel sheet and
#                    updating item quantity in 'Item' and 'Inventory'excel sheet.

# functions: 1. read_from_excel for gettiing item details from 'Item' sheet.
#            2. update_quantity_order for updating-subtracting  item quantity in 'item' sheet
#            3. update_quantity_order for updating- adding item quantity in 'Inventory' sheet.
#            4. get_avail_item for searching functionality used while creating 'normal order'.

#-------------------Importing libraries and clsses-------------------------

import os
import pandas as pd
import openpyxl

class Item:

    # Updating item quantity in 'item' worksheet when order is placed.
    def update_quantity_order(self, store_id, item_name, quantity_ordered, quantity):
        try:
            df = self.read_from_excel()
            temp = df.loc[df['Item Name'] == item_name, ['store id', 'quantity']]
            temp1 = temp.loc[df['store id'] == int(store_id), ['quantity']].to_string()
            detail = temp1.split()
            item_row = detail[1]
            quantity= int(quantity)
            quantity_ordered= int(quantity_ordered)
            quantity -= quantity_ordered
            item_row= int(item_row) + 2
            path = os.getcwd().replace("\\", "/")
            file = path + '/SampleData.xlsx'
            wb = openpyxl.load_workbook(filename=file)
            ws = wb['Item']
            ws.cell(row= item_row,column= 4, value= quantity)
            wb.save(file)
        except PermissionError:
            print("Permission to excel file denied. Please ensure that the file is closed and run the program again")
            exit(0)

    # Updating item quantity in 'Inventory' worksheet when item is added to inventory
    def update_quantity_inv(self, store_id, item_name, quantity_added, quantity):
        df = self.read_from_excel()
        temp = df.loc[df['Item Name'] == item_name, ['store id', 'quantity']]
        temp1 = temp.loc[df['store id'] == int(store_id), ['quantity']].to_string()
        detail = temp1.split()
        item_row = detail[1]
        quantity = int(quantity)
        quantity_added=int(quantity_added)
        quantity += quantity_added
        item_row= int(item_row) + 2
        path = os.getcwd().replace("\\", "/")
        file = path + '/SampleData.xlsx'
        wb = openpyxl.load_workbook(filename=file)
        ws = wb['Item']
        ws.cell(row=item_row, column=4, value=quantity)
        wb.save(file)

    # Following function is created for getting item list on user search while creating normal order
    def get_available_items(self, item_name, store_id):
        searchedItemList=[]
        displayChoice = 1
        df = self.read_from_excel()
        filter_data = df.loc[df['store id'] == store_id, ['Item Name', 'quantity', 'Cost', 'item category']]
        filter_data2= filter_data.loc[df['item category'] != "Raw Material", ['Item Name', 'quantity', 'Cost']]
        filter_data_list = filter_data2.values.tolist()
        if item_name == "roast coffee beans":
            advance_flag = False
        else:
            advance_flag = True
        for data_row in filter_data_list:
            if item_name in data_row[0]:
                if advance_flag and data_row[0].strip() != "roast coffee beans":
                    searchedItemList.append([displayChoice,data_row[0],data_row[1],data_row[2]])
                    displayChoice = displayChoice + 1
                if not advance_flag:
                    searchedItemList.append([displayChoice, data_row[0], data_row[1], data_row[2]])
                    displayChoice = displayChoice + 1
        return searchedItemList

    # Following function is created for getting item name while partial search name
    def get_available_items_inv(self, item_name, store_id):
        searchedItemList=[]
        displayChoice = 1
        df = self.read_from_excel()
        filter_data = df.loc[df['store id'] == store_id, ['Item Name', 'quantity', 'Cost']]
        filter_data_list = filter_data.values.tolist()
        if item_name == "roast coffee beans":
            advance_flag = False
        else:
            advance_flag = True
        for data_row in filter_data_list:
            if item_name in data_row[0]:
                if advance_flag and data_row[0].strip() != "roast coffee beans":
                    searchedItemList.append([displayChoice,data_row[0],data_row[1],data_row[2]])
                    displayChoice = displayChoice + 1
                if not advance_flag:
                    searchedItemList.append([displayChoice, data_row[0], data_row[1], data_row[2]])
                    displayChoice = displayChoice + 1
        return searchedItemList

    def get_quantity(self, item_name, store_id):
        df = self.read_from_excel()
        temp = df.loc[df['Item Name'] == item_name, ['store id', 'quantity']]
        temp1 = temp.loc[df['store id'] == int(store_id), ['quantity']].to_string()
        return temp1.split()[2]

    def get_item_category(self, item_name):
        df = self.read_from_excel()
        temp = df.loc[df['Item Name'] == item_name, ['item category']]
        category = temp.values.tolist()[0]
        return category[0]

    def get_item_id(self, item_name):
        df = self.read_from_excel()
        temp = df.loc[df['Item Name'] == item_name, ['item id']]
        if temp.empty:
            return ""
        else:
            category = temp.values.tolist()[0]
            return category[0]

    def read_from_excel(self):
        path = os.getcwd().replace("\\", "/")
        data = pd.read_excel(path + '/SampleData.xlsx', sheet_name='Item')
        df = pd.DataFrame(data)
        return df

