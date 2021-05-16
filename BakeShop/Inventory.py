#--------------------FIT5136_SEM_02_Team 09  Assignment 4B - Bakeshop System -------------------------------
# Author Names & IDs : Dipal Vyas(31847390) , Aashna Jain(29802911) , Nivedha Balasubramanian(30729963), Kanika Bansal(31030319)
# Start Date : 16th September, 2020
# End Date : 1st November, 2020
#------------------------------------------------------------------------------------------------------------
# Class Description: Inventory class is for reading data from inventory and
#                    writing item data to the inventory excel sheet when inventory is added.

# functions: 1.read_from_excel 2. write_inventory

#-------------------Importing libraries and clsses-------------------------
import os
import pandas as pd
import openpyxl
#--------------------------------------------------------------------------

class Inventory:

# following function opens excel sheet named 'inventory' which is stored in the system using dataframe.
    def read_from_excel(self):
        path = os.getcwd().replace("\\", "/")
        data = pd.read_excel(path + '/SampleData.xlsx', sheet_name='inventory')
        df = pd.DataFrame(data)
        return df

# Following function writes item data which is added to the inventory.
    def write_inventory(self,store_id, item_id, Item_Name, quantity, date_inventory_added):
        try:
            row1 = [store_id, item_id, Item_Name, quantity, date_inventory_added]
            path = os.getcwd().replace("\\", "/")
            file = path + '/SampleData.xlsx'
            wb = openpyxl.load_workbook(filename=file)
            ws = wb['inventory']
            row = ws.max_row + 1
            for col, entry in enumerate(row1, start=1):
                ws.cell(row=row, column=col, value=entry)
            wb.save(file)
        except PermissionError:
            print("Permission to excel file denied. Please ensure that the file is closed and run the program again")
            exit(0)



